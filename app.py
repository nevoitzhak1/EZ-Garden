import firebase_admin
import requests
import asyncio
import json
import openpyxl
from io import BytesIO
from datetime import datetime, timedelta
from flask_mail import Mail, Message
from firebase_admin import credentials, firestore
from flask import Flask, render_template, request, redirect, session, send_from_directory,Response,jsonify
from flask_caching import Cache
from openplantbook_sdk import OpenPlantBookApi, MissingClientIdOrSecret, ValidationError



SENSOR_API_LOGIN_URL = "https://atapi.atomation.net/auth/login"
SENSOR_API_READINGS_URL = "https://atapi.atomation.net/sensors_readings"
OPENPLANT_API_BASE_URL = "https://open.plantbook.io/api/v1"
SENSOR_APP_VERSION = "1.0.0"
SENSOR_ACCESS_TYPE = 5



BASE_URL = "https://open.plantbook.io/api/v1"
CLIENT_ID = "tC0yqwKuCLpFRkFNjTFsHugNJo6poO0I8neJFZuR"
CLIENT_SECRET = "Fny45ACTmtXTOwpC5j3b00HeDgvboHZMh8mjbH190FmeNBNsAWH3mjNm6gYNg38FMyszvzTzhsy2GRdhPb8YcY6uMmD4NYnhhUKPac8a6h1Zgh35IIjDrfRDqYiGmTdT"



app = Flask(__name__)



app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'ezgarden.sce@gmail.com'  
app.config['MAIL_PASSWORD'] = 'pyas pbfo xstu hksu'  
app.config['MAIL_DEFAULT_SENDER'] = 'ezgarden.sce@gmail.com'

mail = Mail(app)



app.config['CACHE_TYPE'] = 'SimpleCache'
app.config['CACHE_DEFAULT_TIMEOUT'] = 3600  
cache = Cache(app)



app.secret_key = "ezgarden"



cred = credentials.Certificate("ez-garden-firebase-adminsdk-rr8ir-37703be9f3.json")
firebase_admin.initialize_app(cred)
db = firestore.client()



@app.context_processor
def inject_user_info():
    """Inject user info into all templates."""
    user_email = session.get("user_email")
    
    return {"user_email": user_email}



@app.route('/')
def home():
    return render_template('home1.html')


#comment 

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')

        if not email or not password:
            return render_template('login.html', error_message="Both email and password are required!")

        employees_ref = db.collection("employees").where("company_email", "==", email).where("password", "==", password).get()
        # if employee
        if employees_ref:
            employee_data = employees_ref[0].to_dict()
            role = employee_data.get("role", 1)  
            company_name = employee_data.get("company_name")

            #Session
            session['user_email'] = email
            session['user_role'] = role
            session['company_name'] = company_name

            if role == 1:  
                return redirect("/home_company_1")
            elif role == 2:   
                return redirect("/home_company_2")
            elif role == 3:  
                return redirect("/home_company_3")

        # if user
        users_ref = db.collection("users").where("email", "==", email).where("password", "==", password).get()
        if users_ref:
            session['user_email'] = email
            session['user_type'] = "private"
            return redirect("/home3")

        return render_template('login.html', error_message="Invalid email or password!")

    return render_template('login.html')



@app.route('/register_type')
def register_type():
    return render_template('register_type.html')



@app.route('/register_private', methods=['GET', 'POST'])
def register_private():
    if request.method == 'POST':
        fullname = request.form.get('fullname')
        email = request.form.get('email')
        password = request.form.get('password')

        if not fullname or not email or not password:
            return render_template('register_private.html', error_message="All fields are required!")

        #Saving data in Firebase
        user_data = {
            "type": "private",
            "fullname": fullname,
            "email": email,
            "password": password
        }
        db.collection("users").add(user_data)
        
        try:
            msg = Message('Welcome to EZ-GARDEN!',
                      recipients=[email])
            msg.body = 'Thank you for registering! We are excited to have you on board.'
            mail.send(msg)
            print("Email sent successfully!")
        except Exception as e:
            print(f"Failed to send email: {e}")

        # success
        return render_template('success.html', message=f"User {fullname} successfully registered with email {email}!")
    
    

    return render_template('register_private.html')



@app.route('/register_company', methods=['GET', 'POST'])
def register_company():
    if request.method == 'POST':
        company_name = request.form.get('company_name')
        name = request.form.get('name')
        company_email = request.form.get('email')
        password = request.form.get('password')

        # Checking if all fields are complete
        if not company_name or not name or not company_email or not password:
            return render_template('register_company.html', error_message="All fields are required!")

        #Saving data in Firebase
        company_data = {
            "company_name": company_name,
            "name": name,
            "company_email": company_email,
            "password": password,
            "role": 3  
        }
        db.collection("employees").add(company_data)
        
        try:
            msg = Message('Welcome to EZ-GARDEN!',
                      recipients=[company_email])
            msg.body = 'Thank you for registering! We are excited to have you on board.'
            mail.send(msg)
            print("Email sent successfully!")
        except Exception as e:
            print(f"Failed to send email: {e}")

        # success 
        success_message = f"The company {company_name} has been successfully registered with the email {company_email}!"
        return render_template('success.html', message=success_message)

    return render_template('register_company.html')



@app.route('/home_company_3')
def home_company_3():
    return render_template('home_company_3.html')



@app.route('/home_company_2')
def home_company_2():
    return render_template('home_company_2.html')



@app.route('/home_company_1')
def home_company_1():
    return render_template('home_company_1.html')



@app.route("/employee_management", methods=["GET", "POST"])
def employee_management():
    company_email = session.get("user_email")
    company_name = session.get("company_name")
    
    # Connected administrator email
    if not company_email:
        print("No company email in session. Redirecting to login.")
        return redirect('/login')  

    if request.method == 'POST':
        # get data
        employee_name = request.form.get('employee_name')
        employee_password = request.form.get('employee_password')
        employee_role = request.form.get('employee_role')

        # chack data
        if not employee_name or not employee_password or not employee_role:
            print("Missing field(s) in the form submission.")
            return render_template(
                'employee_management.html',
                error_message="All fields are required!",
                employees=get_employees_from_db(company_email)
            )

        if not employee_role.isdigit() or int(employee_role) not in [1, 2, 3]:
            print("Invalid role value.")
            return render_template(
                'employee_management.html',
                error_message="Role must be 1, 2, or 3!",
                employees=get_employees_from_db(company_email)
            )

        # Preparing data for Firebase
        employee_data = {
            "company_name": company_name.strip(),
            "company_email": company_email.strip(),
            "name": employee_name.strip(),
            "password": employee_password.strip(),
            "role": int(employee_role)
            
        }

        #Saving data in Firebase
        try:
            print("Attempting to add employee to Firebase:", employee_data)
            db.collection("employees").add(employee_data)
            print("Successfully added employee:", employee_data)
        except Exception as e:
            print("Error adding employee to Firebase:", e)
            return render_template(
                'employee_management.html',
                error_message="Failed to save employee to the database!",
                employees=get_employees_from_db(company_email)
            )

        return redirect('/employee_management')

    # Pulling data from employees with the same email
    try:
        employees = get_employees_from_db(company_email)
        print("Employees retrieved from database:", employees)
    except Exception as e:
        print("Error retrieving employees:", e)
        return render_template(
            'employee_management.html',
            error_message="Failed to retrieve employees!",
            employees=[]
        )

    return render_template('employee_management.html', employees=employees)



def get_employees_from_db(company_email):
    try:
        employees_ref = db.collection("employees").where("company_email", "==", company_email).stream()
        employees = []
        for emp in employees_ref:
            employee_data = emp.to_dict()
            employee_data['id'] = emp.id  # save id
            employees.append(employee_data)
        return employees
    except Exception as e:
        print("Error retrieving employees from Firebase:", e)
        return []



@app.route("/delete_employee/<employee_id>", methods=["POST"])
def delete_employee(employee_id):
    try:
        # Deleting an employee by id
        db.collection("employees").document(employee_id).delete()
        print(f"Employee {employee_id} deleted successfully.")
    except Exception as e:
        print(f"Error deleting employee {employee_id}: {e}")
        return redirect('/employee_management')

    # back
    return redirect('/employee_management')



@app.route('/map')
def map():
    return render_template('map.html')




@app.route('/weather', methods=['GET', 'POST'])
def weather():
    if request.method == 'POST':
        location = request.form.get('location')  # get location

        if not location:
            return "Please provide a location.", 400

        # call API
        url = f"https://api.tomorrow.io/v4/weather/realtime?location={location}&apikey=cAKQ5PlikazuWRCN01CpQnFuzes6rcmt"
        headers = {"accept": "application/json"}

        try:
            response = requests.get(url, headers=headers)
            response.raise_for_status()  
            data = response.json()

            return render_template("weather_data.html", weather_data=data)
        except requests.exceptions.RequestException as e:
            return f"Failed to fetch weather data: {e}", 500

    return render_template("weather.html")



@app.route('/static/map.geojson')
def geojson():
    return send_from_directory('static', 'map.geojson')



@app.route('/neighborhood/<name>')
def neighborhood(name):
    return f"<h1>Welcome to {name.capitalize()} neighborhood</h1>"



@app.route('/home3')
def home3():
    user_email = session.get('user_email')  # get email

    if not user_email:
        return redirect('/login') 

    #Searched on Firebase
    plants_ref = db.collection("plants")
    query = plants_ref.where("user_email", "==", user_email).get()

    #Create a plant list
    plants = [
    {
        "id": plant.id,  # Document ID
        "area": plant.to_dict().get("area", "Unknown"),  
        "type": plant.to_dict().get("type", "Unknown")   
    }
    for plant in query
    ]

    # sending home3.html
    return render_template('home3.html', plants=plants)




@app.route('/export_to_excel', methods=['GET'])
def export_to_excel():
    user_email = session.get('user_email')
    if not user_email:
        return redirect('/login')

    # חיפוש נתוני הצמחים ב-Firebase
    plants_ref = db.collection("plants")
    query = plants_ref.where("user_email", "==", user_email).get()

    # יצירת קובץ Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Plants"

    # הוספת כותרות
    sheet.append(["Area", "Type"])

    # הוספת נתונים
    for plant in query:
        plant_data = plant.to_dict()
        sheet.append([plant_data.get("area", "Unknown"), plant_data.get("type", "Unknown")])

    # שמירה ל-BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # החזרת קובץ למשתמש
    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=plants.xlsx"}
    )


@app.route('/delete_plant/<plant_id>', methods=['POST'])
def delete_plant(plant_id):
    user_email = session.get('user_email')

    if not user_email:
        return redirect('/login')

    db.collection("plants").document(plant_id).delete()

    return redirect('/home3')



@app.route('/add_plant', methods=['GET', 'POST'])
def add_plant():
    user_email = session.get('user_email')  # get email

    if not user_email:
        return redirect('/login') 

    if request.method == 'POST':
        area = request.form.get('area')
        type = request.form.get('type')

        if not area or not type:
            return "<h1 style='text-align:center; color:red;'>All fields are required!</h1>"

        #Saving the plant in Firebase
        plant_data = {
            "area": area,
            "type": type,
            "user_email": user_email
        }
        db.collection("plants").add(plant_data)

        return redirect("/home3")

    return render_template('add_plant.html')



# Access Token OpenPlantBookApi
def get_access_token(client_id, client_secret):
    url = "https://open.plantbook.io/api/v1/token/"
    data = {
        "grant_type": "client_credentials",
        "client_id": 'tC0yqwKuCLpFRkFNjTFsHugNJo6poO0I8neJFZuR',
        "client_secret": 'Fny45ACTmtXTOwpC5j3b00HeDgvboHZMh8mjbH190FmeNBNsAWH3mjNm6gYNg38FMyszvzTzhsy2GRdhPb8YcY6uMmD4NYnhhUKPac8a6h1Zgh35IIjDrfRDqYiGmTdT',
        "scope": "read"
    }

    response = requests.post(url, data=data)

    if response.status_code == 200:
        token_data = response.json()
        return token_data.get("access_token")
    else:
        print(f"Error: Unable to fetch token. Status code: {response.status_code}")
        print(f"Response: {response.text}")
        return None



#Getting information for a plant
async def get_plant_details(plant_type, access_token):
    url = f"https://open.plantbook.io/api/v1/plant/search?alias={plant_type}"
    headers = {"Authorization": f"Bearer {access_token}"}

    try:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            search_results = response.json()
            if search_results and 'results' in search_results and len(search_results['results']) > 0:
                plant = search_results['results'][0]  
                plant_pid = plant['pid']
                
             
                detail_url = f"https://open.plantbook.io/api/v1/plant/detail/{plant_pid}"
                detail_response = requests.get(detail_url, headers=headers)
                
                if detail_response.status_code == 200:
                    return detail_response.json()
                else:
                    print(f"Error fetching plant details: {detail_response.status_code}")
                    return None
            else:
                print("No results found for the plant type.")
                return None
        else:
            print(f"Error fetching plant data: {response.status_code}")
            return None
    except Exception as e:
        print(f"An error occurred: {e}")
        return None



def get_sensor_api_token():

    url = "https://atapi.atomation.net/api/v1/s2s/v1_0/auth/login"
    headers = {
        "accept": "application/json",
        "app_version": "v1.0",
        "access_type": "5",
        "Content-Type": "application/json"
    }
    payload = {
        "email": "sce@atomation.net",
        "password": "123456"
    }

    try:
        response = requests.post(url, headers=headers, data=json.dumps(payload))
        response.raise_for_status()
        
        data = response.json()
        token = data.get("data", {}).get("token")
        if token:
            print("Token successfully retrieved!")
            return token
        else:
            print("Error: Token not found in response.")
            return None
    except requests.exceptions.RequestException as e:
        print(f"Error fetching token: {e}")
        return None



def get_last_sensor_reading(token, mac_address):
 
    url = "https://atapi.atomation.net/api/v1/s2s/v1_0/sensors_readings"

    if not mac_address:
        print("MAC address is missing.")
        return {"error": "MAC address is missing."}

    #Create a date range
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(hours=1)
    formatted_start_date = start_date.strftime("%Y-%m-%dT%H:%M:%S.%fZ")
    formatted_end_date = end_date.strftime("%Y-%m-%dT%H:%M:%S.%fZ")

    print("Start Date:", formatted_start_date)
    print("End Date:", formatted_end_date)


    payload = {
        "filters": {
            "start_date": formatted_start_date,
            "end_date": formatted_end_date,
            "mac": [mac_address],
            "createdAt": True
        },
        "limit": {
            "page": 1,
            "page_size": 1
        }
    }


    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }

    try:
        # Sending to api
        response = requests.post(url, data=json.dumps(payload), headers=headers)
        print(f"Response Status Code: {response.status_code}")

        if response.status_code == 200:
            data = response.json()
            print("Response Data:", data)

            # Check if data exists
            if "data" in data and "readings_data" in data["data"]:
                readings_data = data["data"]["readings_data"]
                if isinstance(readings_data, list) and len(readings_data) > 0:
                    return readings_data[0]  # last reading
                else:
                    print("No data found in the sensor.")
                    return {"message": "No sensor data available."}
            else:
                print("Unexpected data structure from API response.")
                return {"error": "Unexpected data structure from API response."}
        else:
            print(f"Error receiving sensor reading: {response.status_code}")
            print(":", response.text)
            return {"error": f"Sensor API returned status code {response.status_code}"}
    except json.JSONDecodeError:
        print("Response is not a valid JSON")
        return {"error": "Response is not a valid JSON."}
    except Exception as e:
        print(f"General error: {e}")
        return {"error": f"An unexpected error occurred: {e}"}



#OpenPlantBookApi
@app.route('/plant_details/<plant_type>')
def plant_details(plant_type):

    #Access Token
    access_token = get_access_token(CLIENT_ID, CLIENT_SECRET)
    if not access_token:
        return "Error: Unable to retrieve access token.", 500

    details = asyncio.run(get_plant_details(plant_type, access_token))

    #chacking    
    if not details:
        return "Error: Unable to retrieve plant details.", 500
    
    
    
    #atomation
    sensor_token = cache.get("sensor_api_token")
    if not sensor_token:
        # if token empty
        sensor_token = get_sensor_api_token()
        if not sensor_token:
            return "Error: Unable to retrieve sensor token.", 500


        #save token for 1 hour
        cache.set("sensor_api_token", sensor_token, timeout=3600)

    #checking
    sensor_cache_key = "last_sensor_reading"
    sensor_data = cache.get(sensor_cache_key)
    if not sensor_data:
        #get sensor data
        sensor_data = get_last_sensor_reading(sensor_token, "E9:19:79:09:A1:AD")
        if sensor_data:
            cache.set(sensor_cache_key, sensor_data, timeout=60)  # save the last reading for 1 min

    return render_template('plant_details.html', plant=details, sensor=sensor_data)



@app.route('/plants_by_area', methods=['GET'])
def plants_by_area():
    neighborhood = request.args.get('neighborhood')  #  get neighborhood name

    if not neighborhood:
        return "<h1 style='color:red;'>Neighborhood not specified!</h1>"

    user_email = session.get('user_email')  # get email

    if not user_email:
        return redirect('/login')  

    # Search Firebase by neighborhood
    plants_ref = db.collection("plants")
    query = (
        plants_ref
        .where("user_email", "==", user_email)
        .where("area", "==", neighborhood)
        .get()
    )

    # Create a list of plants
    plants = [
        {
            "id": plant.id,  # Document ID
            "area": plant.to_dict().get("area", "Unknown"),  
            "type": plant.to_dict().get("type", "Unknown")   
        }
        for plant in query
    ]

    return render_template('plants_by_area.html', plants=plants, neighborhood=neighborhood)




@app.route('/graphs_select_date', methods=['GET', 'POST'])
def graphs_select_date():
    if request.method == 'POST':
        selected_years = request.form.getlist('year[]')
        selected_months = request.form.getlist('month[]')
        selected_locations = request.form.getlist('locations[]')

        # Format the selected dates and locations
        selected_data = {
            "years": selected_years,
            "months": selected_months,
            "locations": selected_locations
        }

        # Redirect to the graphs page with the data
        return redirect('graphs', data=json.dumps(selected_data))

    return render_template('graphs_select_date.html')


@app.route('/graphs', methods=['GET', 'POST'])
def graphs():
    data = request.args.get('data')
    if data:
        selected_data = json.loads(data)
        years = selected_data.get("years", [])
        months = selected_data.get("months", [])
        locations = selected_data.get("locations", [])
    else:
        years, months, locations = [], [], []

    return render_template('graphs.html', years=years, months=months, locations=locations)


@app.route('/export_graph', methods=['POST'])
def export_graph():
    export_type = request.form.get('type')  # 'excel' or 'pdf'
    graph_data = request.form.get('graph_data')  # Graph data passed as JSON

    if export_type == 'excel':
        # Generate Excel file
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Graph Data"

        # Add headers and data
        graph_dict = json.loads(graph_data)
        headers = list(graph_dict.keys())
        sheet.append(headers)
        rows = zip(*graph_dict.values())
        for row in rows:
            sheet.append(row)

        # Save Excel file to memory
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        return Response(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment;filename=graph_data.xlsx"}
        )

    elif export_type == 'pdf':
        from fpdf import FPDF

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        pdf.cell(200, 10, txt="Graph Data", ln=True, align='C')

        graph_dict = json.loads(graph_data)
        for key, values in graph_dict.items():
            pdf.cell(200, 10, txt=f"{key}: {', '.join(map(str, values))}", ln=True)

        output = BytesIO()
        pdf.output(output)
        output.seek(0)

        return Response(
            output,
            mimetype="application/pdf",
            headers={"Content-Disposition": "attachment;filename=graph_data.pdf"}
        )
  

  
























with open("static/water_consumption.json", "r") as file:
    data_water = json.load(file)
    
    
@app.route("/data_water")
def get_water():
    return jsonify(data_water)



def get_sensor_by_date(token, mac_address, start_date, end_date):
    
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(hours=24)
    formatted_start_date = start_date.strftime("%Y-%m-%dT%H:%M:%S.%fZ")
    formatted_end_date = end_date.strftime("%Y-%m-%dT%H:%M:%S.%fZ")
    url = "https://atapi.atomation.net/api/v1/s2s/v1_0/sensors_readings"

    if not mac_address:
        print("MAC address is missing.")
        return {"error": "MAC address is missing."}

    # Format start and end dates
    formatted_start_date = start_date.strftime("%Y-%m-%dT%H:%M:%S.%fZ")
    formatted_end_date = end_date.strftime("%Y-%m-%dT%H:%M:%S.%fZ")

    print("Start Date:", formatted_start_date)
    print("End Date:", formatted_end_date)

    payload = {
        "filters": {
            "start_date": formatted_start_date,
            "end_date": formatted_end_date,
            "mac": [mac_address],
            "createdAt": True
        },
        "limit": {
            "page": 1,
            "page_size": 1000
        }
    }

    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }

    try:
        # Sending request to API
        response = requests.post(url, data=json.dumps(payload), headers=headers)
        print(f"Response Status Code: {response.status_code}")

        if response.status_code == 200:
            data = response.json()
            print("Response Data (Full):", data)

            # Validate structure
            if "data" in data and "readings_data" in data["data"]:
                readings_data = data["data"]["readings_data"]
                print(f"Number of readings: {len(readings_data)}")
                if isinstance(readings_data, list) and len(readings_data) > 0:
                    return readings_data  # Return all readings
                else:
                    print("No data found in the sensor.")
                    return {"message": "No sensor data available."}
            else:
                print("Unexpected data structure from API response.")
                return {"error": "Unexpected data structure from API response."}
        else:
            print(f"Error receiving sensor reading: {response.status_code}")
            print("Error Response Text:", response.text)
            return {"error": f"Sensor API returned status code {response.status_code}"}
    except json.JSONDecodeError:
        print("Response is not a valid JSON")
        return {"error": "Response is not a valid JSON."}
    except Exception as e:
        print(f"General error: {e}")
        return {"error": f"An unexpected error occurred: {e}"}
    
    
    
    
@app.route("/data_sensor_by_date")
def data_sensor_by_date():
    sensor_token = cache.get("sensor_api_token")
    if not sensor_token:
        # אם אין טוקן, שלוף טוקן חדש
        sensor_token = get_sensor_api_token()
        if not sensor_token:
            return jsonify({"error": "Unable to retrieve sensor token"}), 500

        # שמור את הטוקן ב-Cache לשעה
        cache.set("sensor_api_token", sensor_token, timeout=3600)

    # בדוק אם יש נתונים ב-Cache
    sensor_cache_key = "last_sensor_reading"
    sensor_data = cache.get(sensor_cache_key)
    if not sensor_data:
        # שליפת נתוני החיישן
        start_date = datetime(2024, 12, 10)  # תאריך התחלה לדוגמה
        end_date = datetime(2024, 12, 19)  # תאריך סיום לדוגמה
        print("Requesting sensor data...")
        try:
            sensor_data = get_sensor_by_date(sensor_token, "E9:19:79:09:A1:AD", start_date, end_date)
            if sensor_data:
                # שמור את הנתונים ב-Cache לדקה
                cache.set(sensor_cache_key, sensor_data, timeout=60)
            else:
                print("No sensor data available.")
                return jsonify({"error": "No sensor data available"}), 404
        except Exception as e:
            print(f"Error retrieving sensor data: {e}")
            return jsonify({"error": f"Error retrieving sensor data: {e}"}), 500

    # שליפת טמפרטורה, לחות ותאריך
    try:
        filtered_data = [
            {
                "Temperature": reading["Temperature"],
                "Humidity": reading["Humidity"],
                "SampleTime": reading["sample_time_utc"]
            }
            for reading in sensor_data
        ]
    except KeyError as e:
        print(f"Missing expected key in sensor data: {e}")
        return jsonify({"error": f"Missing key in sensor data: {e}"}), 500

    print("Filtered Sensor Data:", filtered_data)
    return jsonify(filtered_data)



















if __name__ == "__main__":
    app.run(debug=True)
    